import streamlit as st
import openpyxl
import cv2
import numpy as np
import requests
from io import BytesIO
from keras.models import load_model
from openpyxl.styles import PatternFill
from urllib.parse import urlparse
import base64

# Function to check if a URL is valid
def is_valid_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except ValueError:
        return False

# Function to check if an image is a selfie
def is_selfie(image_url, model):
    if not is_valid_url(image_url):
        return False

    response = requests.get(image_url)
    image = cv2.imdecode(np.asarray(bytearray(response.content), dtype="uint8"), cv2.IMREAD_COLOR)
    resized_image = cv2.resize(image, (150, 150))
    resized_image = resized_image / 255.0
    prediction = model.predict(np.array([resized_image]))

    return prediction[0][0] > 0.5

# Function to color non-selfie rows in an Excel file
def color_lines_non_selfie(file_path, model):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the active sheet
    sheet = workbook.active

    # Dynamically find the index of the "Image" column
    image_column_index = None
    for col_num, col_value in enumerate(sheet.iter_cols(max_row=1, values_only=True), start=1):
        if "Image" in col_value:
            image_column_index = col_num
            break

    # Check if the "Image" column was found
    if image_column_index is None:
        st.warning("The 'Image' column was not found in the file.")
        return

    # Iterate over the rows starting from the second row
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        # Check if the image is not a selfie
        if not is_selfie(row[image_column_index - 1].value, model):  # -1 because Python uses 0-based indexing
            # Apply red color fill to the entire row
            for cell in row:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Save the modified Excel file
    modified_file_path = 'modified_file.xlsx'
    workbook.save(modified_file_path)
    st.success(f"Selfie check completed. Check '{modified_file_path}' for results.")

    # Provide download link for the modified file
    st.markdown(get_binary_file_downloader_html(modified_file_path, 'Modified File'), unsafe_allow_html=True)

# Function to create a download link for a file
def get_binary_file_downloader_html(bin_file, file_label='File'):
    with open(bin_file, 'rb') as f:
        data = f.read()
    bin_str = base64.b64encode(data).decode()
    href = f'<a href="data:application/octet-stream;base64,{bin_str}" download="{bin_file}" target="_blank">{file_label}</a>'
    return href

# Streamlit app
def main():
    st.title("Selfie Checker Web App")

    # File upload
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"])

    if uploaded_file is not None:
        try:
            # Load the model
            selfie_model_path = 'selfie_model.h5'
            selfie_model = load_model(selfie_model_path)

            # Color non-selfie rows
            color_lines_non_selfie(uploaded_file, selfie_model)
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")

# Run the app
if __name__ == "__main__":
    main()
